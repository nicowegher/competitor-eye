# --- IMPORTS NECESARIOS ---
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import requests
import io
import os
from datetime import datetime, timedelta
import firebase_admin
from firebase_admin import credentials, firestore, auth as firebase_auth
from google.cloud import storage
from dotenv import load_dotenv
import json
from google.oauth2 import service_account
import threading
import logging
from mailersend import emails
from time import sleep
import mercadopago
import time
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import openpyxl

# --- CONFIGURACIÓN DE LOGGING ---
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# --- CONFIGURACIÓN FLASK Y CORS ---
app = Flask(__name__)
CORS(app)  # CRÍTICO para Firebase Studio

# --- Cargar variables de entorno ---
load_dotenv()
GCS_BUCKET_NAME = os.environ.get('GCS_BUCKET_NAME')
GOOGLE_APPLICATION_CREDENTIALS = os.environ.get('GOOGLE_APPLICATION_CREDENTIALS')

# --- Inicializar Firebase Admin SDK ---
# Lee el JSON de la variable de entorno
firebase_creds = os.environ.get("FIREBASE_SERVICE_ACCOUNT")
if firebase_creds:
    cred_dict = json.loads(firebase_creds)
    cred = credentials.Certificate(cred_dict)
    firebase_admin.initialize_app(cred)
    gcs_credentials = service_account.Credentials.from_service_account_info(cred_dict)
    storage_client = storage.Client(credentials=gcs_credentials, project=cred_dict.get("project_id"))
else:
    # Modo local, usa el archivo
    cred = credentials.Certificate("firebase_service_account.json")
    firebase_admin.initialize_app(cred)
    storage_client = storage.Client()

db = firestore.client()
bucket = storage_client.bucket(GCS_BUCKET_NAME)

# --- VARIABLE GLOBAL PARA EL ESTADO DEL SCRAPER (SIMPLE) ---
scraper_status = {
    "is_running": False,
    "progress": 0,
    "total_tasks": 0,
    "completed_tasks": 0,
    "error": None,
    "result": None,
    "current_user": None
}

# Variable global para controlar si hay un scraper corriendo
scraper_en_proceso = threading.Event()

# --- LÍMITES POR PLAN ---
PLAN_LIMITS = {
    "free_trial": {
        "max_groups": 1,
        "max_competitors": 2,
        "max_days": 7,
        "scheduling": False
    },
    "esencial": {
        "max_groups": 1,
        "max_competitors": 3,
        "max_days": 30,
        "scheduling": False
    },
    "pro": {
        "max_groups": 3,
        "max_competitors": 5,
        "max_days": 60,
        "scheduling": "weekly"
    },
    "market_leader": {
        "max_groups": 5,
        "max_competitors": 7,
        "max_days": 90,
        "scheduling": "weekly"
    },
    "custom": {
        "max_groups": 8,
        "max_competitors": 7,
        "max_days": 90,
        "scheduling": "weekly"
    }
}

# --- Mapeo de planes a preapproval_plan_id de Mercado Pago ---
MP_PLAN_IDS = {
    "esencial": "2c93808497c462520197d744586508be",
    "pro": "2c93808497c19ac40197d7445b440a20",
    "market_leader": "2c93808497d635430197d7445e1c00bc"
}

def get_user_plan(uid):
    try:
        user_ref = db.collection('users').document(uid)
        user_doc = user_ref.get()
        if user_doc is not None and hasattr(user_doc, 'exists') and user_doc.exists:
            user_dict = user_doc.to_dict() if hasattr(user_doc, 'to_dict') else None
            if user_dict is not None:
                return user_dict.get('plan', 'free_trial')
        return 'free_trial'
    except Exception as e:
        logger.error(f"Error obteniendo plan del usuario {uid}: {e}")
        return 'free_trial'

# --- FUNCIÓN ASÍNCRONA PARA EL SCRAPER (SIMPLE) ---
def run_scraper_async(hotel_base_urls, days, userEmail=None, setName=None, nights=1, currency="USD", report_id=None, userId=None, setId=None):
    global scraper_status
    try:
        logger.info(f"[Scraper] INICIO run_scraper_async para reporte: {report_id} | hoteles: {hotel_base_urls}")
        scraper_status["is_running"] = True
        scraper_status["error"] = None
        scraper_status["progress"] = 0
        
        from apify_scraper import scrape_booking_data
        
        logger.info(f"[Scraper] Ejecutando scraper para {len(hotel_base_urls)} hoteles por {days} días, {nights} noches, moneda {currency}")
        result = scrape_booking_data(hotel_base_urls, days, nights, currency)
        
        if not result:
            logger.error(f"[Scraper] ERROR: No se obtuvieron datos del scraper. Antes de raise Exception...")
            raise Exception("No se obtuvieron datos del scraper")
        
        logger.info(f"[Scraper] Datos obtenidos del scraper: {len(result)} hoteles")
        
        # --- ENRIQUECER DATOS PARA GRÁFICOS ---
        hotelNames = [hotel["Hotel Name"] for hotel in result]
        all_dates = set()
        for hotel in result:
            for k in hotel.keys():
                if k not in ("Hotel Name", "URL"):
                    all_dates.add(k)
        # Ordenar fechas cronológicamente de menor a mayor
        from datetime import datetime as dt
        all_dates = sorted(all_dates, key=lambda x: dt.strptime(x, "%Y-%m-%d"))
        column_order = ["Hotel Name", "URL"] + all_dates
        
        logger.info(f"[Scraper] Procesando {len(all_dates)} fechas para {len(hotelNames)} hoteles")
        
        # Generar chartData con el formato correcto para el frontend
        chartData = []
        for date in all_dates:
            day_obj = {"date": date}
            
            # Añadir precios de cada hotel
            for hotel in result:
                name = hotel["Hotel Name"]
                price = hotel.get(date, None)
                if price is not None:
                    try:
                        day_obj[name] = float(price)
                    except:
                        day_obj[name] = None
                else:
                    day_obj[name] = None
            
            # Calcular métricas para esta fecha
            precios_validos = {}
            for hotel in result:
                name = hotel["Hotel Name"]
                price = hotel.get(date, None)
                if price is not None:
                    try:
                        precios_validos[name] = float(price)
                    except:
                        pass
            
            # Calcular promedio de competidores
            hotel_principal = hotelNames[0] if hotelNames else None
            competidores = hotelNames[1:] if len(hotelNames) > 1 else []
            
            if competidores and hotel_principal:
                precios_competidores = [precios_validos.get(comp, 0) for comp in competidores if precios_validos.get(comp, 0) > 0]
                if precios_competidores:
                    promedio = sum(precios_competidores) / len(precios_competidores)
                    day_obj["Tarifa promedio de competidores"] = round(promedio, 2)
                else:
                    day_obj["Tarifa promedio de competidores"] = None
                
                # Calcular disponibilidad (%)
                total_hoteles = len(hotelNames)
                hoteles_con_precio = len([name for name in hotelNames if precios_validos.get(name) is not None])
                disponibilidad_porcentaje = round((hoteles_con_precio / total_hoteles) * 100) if total_hoteles > 0 else 0
                day_obj["Disponibilidad de la oferta (%)"] = disponibilidad_porcentaje
            else:
                day_obj["Tarifa promedio de competidores"] = None
                day_obj["Disponibilidad de la oferta (%)"] = None
            
            chartData.append(day_obj)
        
        logger.info(f"[Scraper] ChartData generado con {len(chartData)} días")
        
        # Asegurar que hotelNames esté correctamente ordenado (hotel principal primero)
        # El primer hotel en la lista debe ser el hotel principal del usuario
        if hotelNames:
            hotel_principal = hotelNames[0]
            competidores = hotelNames[1:]
            # Reordenar si es necesario para asegurar que el hotel principal esté primero
            hotelNames = [hotel_principal] + competidores
        
        # Generar filas de métricas para los archivos Excel/CSV
        promedio_competidores_row = {"Hotel Name": "Tarifa promedio de competidores", "URL": ""}
        disponibilidad_row = {"Hotel Name": "Disponibilidad de la oferta (%)", "URL": ""}
        diferencia_row = {"Hotel Name": "Diferencia de mi tarifa vs. la tarifa promedio de los competidores (%)", "URL": ""}
        
        for date in all_dates:
            precios_validos = {}
            for hotel in result:
                name = hotel["Hotel Name"]
                price = hotel.get(date, None)
                if price is not None:
                    try:
                        precios_validos[name] = float(price)
                    except:
                        pass
            
            if competidores and hotel_principal:
                precios_competidores = [precios_validos.get(comp, 0) for comp in competidores if precios_validos.get(comp, 0) > 0]
                if precios_competidores:
                    promedio = sum(precios_competidores) / len(precios_competidores)
                    promedio_competidores_row[date] = str(round(promedio, 2))
                else:
                    promedio_competidores_row[date] = ''
                
                if hotel_principal in precios_validos:
                    disponibilidad_row[date] = "100"
                else:
                    disponibilidad_row[date] = "0"
                
                if hotel_principal in precios_validos and promedio_competidores_row[date] not in (None, ""):
                    mi_precio = precios_validos[hotel_principal]
                    diff_percent = ((mi_precio - float(promedio_competidores_row[date])) / float(promedio_competidores_row[date])) * 100
                    diferencia_row[date] = str(round(diff_percent, 0))
                else:
                    diferencia_row[date] = ''
            else:
                promedio_competidores_row[date] = ''
                disponibilidad_row[date] = ''
                diferencia_row[date] = ''
        
        result.append(promedio_competidores_row)
        result.append(disponibilidad_row)
        result.append(diferencia_row)
        
        logger.info(f"[Scraper] Métricas calculadas y agregadas al resultado")
        
        # --- GENERAR ARCHIVOS ---
        def limpiar_nombre(nombre):
            import re
            return re.sub(r'[<>:"/\\|?*]', '', nombre)
        set_name_clean = limpiar_nombre(setName) if setName else "scraping"
        if not report_id:
            report_id = f"{set_name_clean}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        # Generar nombres de archivos con formato HotelRateShopper_YYMMDD_NombreSetCompetitivo
        fecha_formato = datetime.now().strftime('%y%m%d')
        nombre_archivo = f"HotelRateShopper_{fecha_formato}_{set_name_clean}"
        csv_blob_name = f"reports/{nombre_archivo}.csv"
        excel_blob_name = f"reports/{nombre_archivo}.xlsx"
        
        logger.info(f"[Scraper] Generando archivos CSV y Excel para report_id: {report_id}")
        
        # Crear DataFrame y forzar orden de columnas
        df_csv = pd.DataFrame(result)
        df_csv = df_csv.reindex(columns=column_order)
        csv_buffer = io.StringIO()
        df_csv.to_csv(csv_buffer, index=False, decimal=',')
        csv_buffer.seek(0)
        csv_blob = bucket.blob(csv_blob_name)
        # Añadir metadatos personalizados con userId para las reglas de seguridad
        csv_blob.metadata = {'userId': userId}
        csv_blob.upload_from_string(csv_buffer.getvalue(), content_type='text/csv')
        logger.info(f"[Scraper] Archivo CSV generado y subido: {csv_blob_name}")
        df_excel = pd.DataFrame(result)
        df_excel = df_excel.reindex(columns=column_order)
        
        # Limpiar nombres de hoteles (eliminar prefijo Ar/)
        def limpiar_nombre_hotel(nombre):
            if nombre and isinstance(nombre, str):
                return nombre.replace('Ar/', '').replace('ar/', '')
            return nombre
        
        # Aplicar limpieza a nombres de hoteles
        df_excel['Hotel Name'] = df_excel['Hotel Name'].apply(limpiar_nombre_hotel)
        
        # Convertir valores numéricos a formato de moneda con coma como separador decimal
        for col in df_excel.columns:
            if col not in ["Hotel Name", "URL"]:  # Solo columnas de fechas
                df_excel[col] = df_excel[col].apply(lambda x: f"{float(x):.2f}".replace('.', ',') if pd.notna(x) and x != '' and str(x).replace('.', '').replace(',', '').isdigit() else x)
        
        # Guardar Excel con pie de página y ajuste de columnas
        from openpyxl import load_workbook
        df_excel.to_excel('temp_excel_upload.xlsx', sheet_name='Tarifas', index=False)
        wb = load_workbook('temp_excel_upload.xlsx')
        ws = wb['Tarifas']
        # --- Formato profesional (colores, título, encabezados, etc) ---
        # Definir colores de la marca
        titulo_fondo = PatternFill(start_color='002A80', end_color='002A80', fill_type='solid')
        titulo_fuente = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        encabezado_fondo = PatternFill(start_color='002A80', end_color='002A80', fill_type='solid')
        encabezado_fuente = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        hotel_principal_fondo = PatternFill(start_color='F0FAFB', end_color='F0FAFB', fill_type='solid')
        hotel_principal_fuente = Font(name='Calibri', size=10, bold=True, color='46B1DE')
        competidor_fuente = Font(name='Calibri', size=10, color='000000')
        metricas_fondo = PatternFill(start_color='F0FAFB', end_color='F0FAFB', fill_type='solid')
        metricas_fuente = Font(name='Calibri', size=10, bold=True, color='000000')
        verde_fuente = Font(name='Calibri', size=10, color='28A745')
        rojo_fuente = Font(name='Calibri', size=10, color='DC3545')
        borde_gris = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        # 1. INSERTAR TÍTULO (Fila 1)
        ws.insert_rows(1)
        titulo_cell = ws['A1']
        titulo_cell.value = "Hotel Rate Shopper"
        titulo_cell.fill = titulo_fondo
        titulo_cell.font = titulo_fuente
        titulo_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Combinar celdas del título
        ultima_columna = get_column_letter(len(df_excel.columns))
        ws.merge_cells(f'A1:{ultima_columna}1')
        
        # Ajustar altura de la fila del título
        ws.row_dimensions[1].height = 30
        
        # 2. ESCRIBIR DATOS (empezando desde fila 3)
        for idx, row in enumerate(df_excel.itertuples(), start=3):
            for col_idx, value in enumerate(row[1:], start=1):
                cell = ws.cell(row=idx, column=col_idx, value=value)
                cell.border = borde_gris
                
                # Aplicar formato según el tipo de fila
                hotel_name = row[1] if len(row) > 1 else ""
                
                if idx == 3:  # Hotel principal
                    cell.fill = hotel_principal_fondo
                    cell.font = hotel_principal_fuente
                elif "Tarifa promedio de competidores" in str(hotel_name) or "Disponibilidad de la oferta" in str(hotel_name):
                    cell.fill = metricas_fondo
                    cell.font = metricas_fuente
                elif "Diferencia de mi tarifa" in str(hotel_name):
                    cell.fill = metricas_fondo
                    cell.font = metricas_fuente
                    # Formato condicional para diferencias
                    if col_idx > 2 and value and str(value).replace('.', '').replace(',', '').replace('-', '').isdigit():
                        try:
                            diff_value = float(str(value).replace(',', '.'))
                            if diff_value < 0:
                                cell.font = verde_fuente
                            elif diff_value > 0:
                                cell.font = rojo_fuente
                        except (ValueError, TypeError):
                            pass
                else:  # Competidores
                    cell.font = competidor_fuente
        
        # 3. FORMATEAR ENCABEZADOS (Fila 2)
        for col_idx, col_name in enumerate(df_excel.columns, start=1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.fill = encabezado_fondo
            cell.font = encabezado_fuente
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = borde_gris
            
            # Borde inferior blanco para encabezados
            cell.border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thick', color='FFFFFF')
            )
        
        # 4. AJUSTAR ANCHO DE COLUMNAS
        for col_idx, col_name in enumerate(df_excel.columns, start=1):
            column_letter = get_column_letter(col_idx)
            if col_name == "Hotel Name":
                ws.column_dimensions[column_letter].width = 25
            elif col_name == "URL":
                ws.column_dimensions[column_letter].width = 50
            else:
                ws.column_dimensions[column_letter].width = 15
        
        # 5. CONGELAR PANELES (Título y encabezados)
        ws.freeze_panes = 'A3'
        
        # 6. ALINEACIÓN DE TEXTO
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for col_idx, cell in enumerate(row, start=1):
                if col_idx <= 2:  # Hotel Name y URL
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:  # Fechas y precios
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # 7. AGREGAR PIE DE PÁGINA
        # Agregar fila vacía después de los datos
        ws.insert_rows(ws.max_row + 1)
        
        # Agregar fila de pie de página
        pie_fila = ws.max_row + 1
        pie_cell = ws.cell(row=pie_fila, column=1, value="Con tecnología de www.HotelRateShopper.com")
        pie_cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        pie_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Pintar fondo azul en toda la fila del pie de página
        for col_idx in range(1, len(df_excel.columns) + 1):
            cell = ws.cell(row=pie_fila, column=col_idx)
            cell.fill = titulo_fondo  # Mismo fondo azul que el título
        
        # 8. GUARDAR EL ARCHIVO CON FORMATO APLICADO
        wb.save('temp_excel_upload.xlsx')
        
        with open('temp_excel_upload.xlsx', 'rb') as f:
            excel_blob = bucket.blob(excel_blob_name)
            # Añadir metadatos personalizados con userId para las reglas de seguridad
            excel_blob.metadata = {'userId': userId}
            excel_blob.upload_from_file(f, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        os.remove('temp_excel_upload.xlsx')
        logger.info(f"[Scraper] Archivo Excel generado y subido: {excel_blob_name}")
        
        # --- GUARDAR EN FIRESTORE ---
        logger.info(f"[Scraper] PREPARANDO para guardar reporte en Firestore con ID: {report_id}")
        now = datetime.now()

        # Generar URLs firmadas para 7 días (máximo permitido por Google Cloud Storage)
        csv_signed_url = csv_blob.generate_signed_url(
            version="v4",
            expiration=timedelta(days=7),
            method="GET"
        )
        excel_signed_url = excel_blob.generate_signed_url(
            version="v4",
            expiration=timedelta(days=7),
            method="GET"
        )
        logger.info(f"[Scraper] URL firmada CSV: {csv_signed_url}")
        logger.info(f"[Scraper] URL firmada Excel: {excel_signed_url}")

        report_data = {
            "status": "completed",
            "csvFileUrl": csv_signed_url,
            "xlsxFileUrl": excel_signed_url,
            "createdAt": now,
            "completedAt": now,
            "chartData": chartData,
            "hotelNames": hotelNames,
            "userId": userId,
            "setId": setId if setId else report_id,
            "setName": setName,
            "result": result,
            "days": days,
            "nights": nights,
            "currency": currency
        }
        
        logger.info(f"[Scraper] ANTES de intentar guardar en Firestore. report_data keys: {list(report_data.keys())}")
        
        try:
            logger.info(f"[Scraper] Guardando documento en Firestore con .set()... (ID: {report_id})")
            db.collection("scraping_reports").document(report_id).set(report_data)
            logger.info(f"[Scraper] ✅ Documento guardado exitosamente en Firestore (ID: {report_id})")
        except Exception as e:
            logger.error(f"[Scraper] ❌ ERROR al guardar documento en Firestore (ID: {report_id}): {e}")
            logger.error(f"[Scraper] ❌ Tipo de error: {type(e)}")
            logger.error(f"[Scraper] ❌ Detalles completos del error: {str(e)}")
        
        # --- ENVIAR CORREO AL USUARIO ---
        try:
            if userEmail:
                logger.info(f"[Scraper] Enviando correo a: {userEmail}")
                from mailersend import emails
                mailer = emails.NewEmail(os.environ.get('MAILERSEND_API_KEY'))
                
                # Definir un diccionario vacío para poblar con valores del correo
                mail_body = {}
                
                # Configurar remitente
                mail_from = {
                    "email": os.environ.get('MAILERSEND_SENDER_EMAIL', 'noreply@hotelrateshopper.com'),
                    "name": os.environ.get('MAILERSEND_SENDER_NAME', 'Hotel Rate Shopper')
                }
                
                # Configurar destinatario
                recipients = [
                    {
                        "email": userEmail,
                        "name": "Usuario"
                    }
                ]
                
                # Configurar reply-to
                reply_to = {
                    "email": "nicolas.wegher@gmail.com",
                    "name": "Nicolás Wegher"
                }
                
                # Configurar el correo usando los métodos de la librería
                mailer.set_mail_from(mail_from, mail_body)
                mailer.set_mail_to(recipients, mail_body)
                mailer.set_subject(f"¡Tu informe para '{setName}' está listo!", mail_body)
                mailer.set_reply_to(reply_to, mail_body)
                mailer.set_html_content(f"""
<!DOCTYPE html>
<html lang="es">
  <head>
    <meta charset="UTF-8">
    <title>¡Tu informe para '{setName}' está listo!</title>
  </head>
  <body style="font-family: Arial, sans-serif; background: #fff; color: #222; margin: 0; padding: 0;">
    <div style="max-width: 600px; margin: 40px auto; padding: 32px 24px; background: #fff; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.04);">
      <h1 style="color: #222; font-size: 2em; margin-bottom: 0.5em;">¡Tu informe está listo!</h1>
      <p>Hola,</p>
      <p>
        Tu informe de precios para el grupo competitivo <strong>'{setName}'</strong> ya está disponible.
      </p>
      <div style="margin: 32px 0;">
        <a href="https://hotelrateshopper.com/" style="display: inline-block; background: #34a853; color: #fff; font-weight: bold; text-decoration: none; padding: 16px 32px; border-radius: 6px; font-size: 1.1em;">
          Ver mi informe en la plataforma
        </a>
      </div>
      <p>También puedes descargar el informe directamente desde este correo. <b>Los enlaces estarán disponibles por 7 días</b>. Luego de ese plazo, siempre podrás acceder a tus informes desde la app.</p>
      <div style="margin: 20px 0;">
        <a href="{excel_signed_url}" style="display: inline-block; background: #4285f4; color: #fff; font-weight: bold; text-decoration: none; padding: 12px 24px; border-radius: 6px; font-size: 1em; margin-bottom: 8px;">
          Descargar Informe (Excel)
        </a>
      </div>
      <div style="margin-bottom: 24px;">
        <a href="{csv_signed_url}" style="color: #4285f4; text-decoration: underline; font-size: 1em;">
          Descargar en formato .CSV
        </a>
      </div>
      <p style="margin-top: 32px;">Gracias por usar <a href="https://hotelrateshopper.com" style="color: #4285f4; text-decoration: underline;">HotelRateShopper.com</a></p>
    </div>
  </body>
</html>
""", mail_body)
                
                response = mailer.send(mail_body)
                logger.info(f"[Scraper] ✅ Correo enviado. Respuesta de MailerSend: {response}")
        except Exception as e:
            logger.error(f"[Scraper] ❌ Error enviando correo: {e}")
        
        # Actualizar estado
        logger.info(f"[Scraper] Actualizando scraper_status...")
        scraper_status["result"] = report_data
        scraper_status["progress"] = 100
        scraper_status["is_running"] = False
        logger.info(f"[Scraper] ✅ FIN run_scraper_async (éxito) - report_id: {report_id}")
        
    except Exception as e:
        logger.error(f"[Scraper] ❌ ERROR en scraper: {e}")
        logger.error(f"[Scraper] ❌ Tipo de error: {type(e)}")
        logger.error(f"[Scraper] ❌ Antes de intentar actualizar documento a failed...")
        
        # Actualizar el documento con status failed y completedAt
        try:
            now = datetime.now()
            logger.info(f"[Scraper] Intentando actualizar documento a failed en Firestore (ID: {report_id})...")
            db.collection("scraping_reports").document(report_id).update({
                "status": "failed",
                "completedAt": now,
                "error": str(e)
            })
            logger.info(f"[Scraper] ✅ Reporte {report_id} marcado como failed en Firestore")
        except Exception as e2:
            logger.error(f"[Scraper] ❌ ERROR actualizando status failed en Firestore: {e2}")
            logger.error(f"[Scraper] ❌ Tipo de error al actualizar: {type(e2)}")
        
        logger.info(f"[Scraper] Actualizando scraper_status con error...")
        scraper_status["error"] = str(e)
        scraper_status["is_running"] = False
        scraper_status["current_user"] = None
        logger.info(f"[Scraper] ❌ FIN run_scraper_async (fallo) - report_id: {report_id}")

def cola_procesadora_scraping():
    global scraper_en_proceso
    while True:
        try:
            if scraper_en_proceso.is_set():
                time.sleep(5)
                continue
            logger.info("[ColaScraping] Bucle activo. Buscando tareas encoladas...")
            query = (
                db.collection('scraping_reports')
                .where('status', '==', 'queued')
                .order_by('createdAt')
                .limit(1)
            )
            docs = list(query.stream())
            if not docs:
                logger.info("[ColaScraping] No se encontraron tareas encoladas. Esperando 20s...")
                time.sleep(20)
                continue
            doc = docs[0]
            doc_ref = doc.reference
            data = doc.to_dict()
            if data is None:
                logger.error(f"[ColaScraping] El documento {doc_ref.id} no tiene datos. Saltando...")
                continue
            # LOGS DE DEPURACIÓN
            logger.info(f"[ColaScraping][DEBUG] Documento Firestore data: {data}")
            logger.info(f"[ColaScraping][DEBUG] ownHotelUrl: {data.get('ownHotelUrl')} (tipo: {type(data.get('ownHotelUrl'))})")
            logger.info(f"[ColaScraping][DEBUG] competitorHotelUrls: {data.get('competitorHotelUrls')} (tipo: {type(data.get('competitorHotelUrls'))})")
            # Obtener hoteles directamente del documento (denormalización)
            hotel_base_urls = []
            if data.get('ownHotelUrl'):
                hotel_base_urls.append(data['ownHotelUrl'])
            comp_urls = data.get('competitorHotelUrls')
            if comp_urls:
                if isinstance(comp_urls, list):
                    hotel_base_urls.extend(comp_urls)
                elif isinstance(comp_urls, str):
                    hotel_base_urls.append(comp_urls)
                else:
                    logger.warning(f"[ColaScraping][DEBUG] competitorHotelUrls tiene un tipo inesperado: {type(comp_urls)}")
            logger.info(f"[ColaScraping][DEBUG] hotel_base_urls final: {hotel_base_urls}")
            if not hotel_base_urls:
                logger.error(f"[ColaScraping] La tarea {doc_ref.id} no tiene hoteles para analizar. Saltando...")
                doc_ref.update({'status': 'failed', 'error': 'No se encontraron hoteles para analizar'})
                continue
            logger.info(f"[ColaScraping] Procesando tarea: {doc_ref.id} - {data.get('setName', '')} con hoteles: {hotel_base_urls}")
            doc_ref.update({'status': 'pending'})
            # Marcar que hay un scraper en proceso
            scraper_en_proceso.set()
            # Ejecutar el scraper con los datos del documento
            run_scraper_async(
                hotel_base_urls,
                data.get('days', data.get('daysToScrape', 7)),
                data.get('userEmail', None),
                data.get('setName', None),
                data.get('nights', 1),
                data.get('currency', 'USD'),
                report_id=doc_ref.id,
                userId=data.get('userId', None),
                setId=data.get('setId', None)
            )
            # Cuando termine, limpiar el flag
            scraper_en_proceso.clear()
        except Exception as e:
            logger.error(f"[ColaScraping] Error en cola_procesadora_scraping: {e}")
        time.sleep(5)

# --- ENDPOINT PRINCIPAL (SIMPLE) ---
@app.route('/run-scraper', methods=['POST'])
def run_scraper():
    global scraper_status
    
    try:
        data = request.get_json()
        uid = data.get('uid')
        hotel_base_urls = data.get('hotel_base_urls', [])
        days = data.get('days', 7)
        nights = data.get('nights', 1)
        currency = data.get('currency', 'USD')
        setName = data.get('setName', 'Mi Set')
        userEmail = data.get('userEmail')
        report_id = data.get('report_id')
        setId = data.get('setId')  # <-- Tomar el setId del payload
        
        logger.info(f"[run-scraper] Recibido UID: {uid}, report_id: {report_id}, setId: {setId}")
        
        # Verificar si ya hay un scraper corriendo
        if scraper_status["is_running"]:
            return jsonify({
                "success": False,
                "message": "Ya hay un scraper ejecutándose. Espera a que termine."
            }), 400
        
        # Verificar plan del usuario
        user_plan = get_user_plan(uid)
        plan_limits = PLAN_LIMITS.get(user_plan, PLAN_LIMITS["free_trial"])
        
        if len(hotel_base_urls) > plan_limits["max_competitors"] + 1:  # +1 para el hotel principal
            return jsonify({
                "success": False,
                "message": f"Tu plan {user_plan} permite máximo {plan_limits['max_competitors']} competidores"
            }), 400
        
        if days > plan_limits["max_days"]:
            return jsonify({
                "success": False,
                "message": f"Tu plan {user_plan} permite máximo {plan_limits['max_days']} días"
            }), 400
        
        # Iniciar scraper en thread separado
        scraper_status["current_user"] = uid
        thread = threading.Thread(
            target=run_scraper_async,
            args=(hotel_base_urls, days, userEmail, setName, nights, currency, report_id, uid, setId) # Pasar userId y setId
        )
        thread.daemon = True
        thread.start()
        
        return jsonify({
            "success": True,
            "message": "Scraper iniciado correctamente",
            "plan": user_plan
        })
        
    except Exception as e:
        logger.error(f"Error en run-scraper: {e}")
        return jsonify({
            "success": False,
            "message": f"Error: {str(e)}"
        }), 500

# --- ENDPOINTS ADICIONALES (MANTENER) ---
@app.route('/', methods=['GET'])
def health_check():
    return jsonify({"status": "ok", "message": "Backend funcionando"})

@app.route('/datos', methods=['GET'])
def obtener_datos():
    try:
        datos = obtener_datos_externos()
        return jsonify(datos)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/descargar-csv', methods=['GET', 'POST'])
def descargar_csv():
    try:
        data = request.get_json() if request.method == 'POST' else request.args
        report_id = data.get('report_id')
        
        if not report_id:
            return jsonify({"error": "report_id requerido"}), 400
        
        # Obtener reporte de Firestore
        report_ref = db.collection("scraping_reports").document(report_id)
        report = report_ref.get()
        
        if not report.exists:
            return jsonify({"error": "Reporte no encontrado"}), 404
        
        report_data = report.to_dict() or {}
        result = report_data.get('result', [])
        
        # Crear CSV
        df = pd.DataFrame(result)
        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False, decimal=',')
        csv_buffer.seek(0)
        
        # Generar nombre de archivo con formato HotelRateShopper_YYMMDD_NombreSetCompetitivo
        fecha_formato = datetime.now().strftime('%y%m%d')
        nombre_archivo = f"HotelRateShopper_{fecha_formato}_{report_data.get('setName', 'Reporte')}.csv"
        
        return send_file(
            io.BytesIO(csv_buffer.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True, 
            download_name=nombre_archivo
        )
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/descargar-excel', methods=['GET', 'POST'])
def descargar_excel():
    try:
        data = request.get_json() if request.method == 'POST' else request.args
        report_id = data.get('report_id')
        
        if not report_id:
            return jsonify({"error": "report_id requerido"}), 400
        
        # Obtener reporte de Firestore
        report_ref = db.collection("scraping_reports").document(report_id)
        report = report_ref.get()
        
        if not report.exists:
            return jsonify({"error": "Reporte no encontrado"}), 404
        
        report_data = report.to_dict() or {}
        result = report_data.get('result', [])
        
        # Crear Excel
        df = pd.DataFrame(result)
        
        # Limpiar nombres de hoteles (eliminar prefijo Ar/)
        def limpiar_nombre_hotel(nombre):
            if nombre and isinstance(nombre, str):
                return nombre.replace('Ar/', '').replace('ar/', '')
            return nombre
        
        # Aplicar limpieza a nombres de hoteles
        df['Hotel Name'] = df['Hotel Name'].apply(limpiar_nombre_hotel)
        
        # Convertir valores numéricos a formato de moneda con coma como separador decimal
        for col in df.columns:
            if col not in ["Hotel Name", "URL"]:  # Solo columnas de fechas
                df[col] = df[col].apply(lambda x: f"{float(x):.2f}".replace('.', ',') if pd.notna(x) and x != '' and str(x).replace('.', '').replace(',', '').isdigit() else x)
        
        # Guardar Excel con pie de página y ajuste de columnas (descarga)
        from openpyxl import load_workbook
        df.to_excel('temp_excel_download.xlsx', sheet_name='Tarifas', index=False)
        wb = load_workbook('temp_excel_download.xlsx')
        ws = wb['Tarifas']
        # --- Formato profesional (colores, título, encabezados, etc) ---
        # Definir colores de la marca
        titulo_fondo = PatternFill(start_color='002A80', end_color='002A80', fill_type='solid')
        titulo_fuente = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        encabezado_fondo = PatternFill(start_color='002A80', end_color='002A80', fill_type='solid')
        encabezado_fuente = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        hotel_principal_fondo = PatternFill(start_color='F0FAFB', end_color='F0FAFB', fill_type='solid')
        hotel_principal_fuente = Font(name='Calibri', size=10, bold=True, color='46B1DE')
        competidor_fuente = Font(name='Calibri', size=10, color='000000')
        metricas_fondo = PatternFill(start_color='F0FAFB', end_color='F0FAFB', fill_type='solid')
        metricas_fuente = Font(name='Calibri', size=10, bold=True, color='000000')
        verde_fuente = Font(name='Calibri', size=10, color='28A745')
        rojo_fuente = Font(name='Calibri', size=10, color='DC3545')
        borde_gris = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        # 1. INSERTAR TÍTULO (Fila 1)
        ws.insert_rows(1)
        titulo_cell = ws['A1']
        titulo_cell.value = "Hotel Rate Shopper"
        titulo_cell.fill = titulo_fondo
        titulo_cell.font = titulo_fuente
        titulo_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Combinar celdas del título
        ultima_columna = get_column_letter(len(df.columns))
        ws.merge_cells(f'A1:{ultima_columna}1')
        
        # Ajustar altura de la fila del título
        ws.row_dimensions[1].height = 30
        
        # 2. ESCRIBIR DATOS (empezando desde fila 3)
        for idx, row in enumerate(df.itertuples(), start=3):
            for col_idx, value in enumerate(row[1:], start=1):
                cell = ws.cell(row=idx, column=col_idx, value=value)
                cell.border = borde_gris
                
                # Aplicar formato según el tipo de fila
                hotel_name = row[1] if len(row) > 1 else ""
                
                if idx == 3:  # Hotel principal
                    cell.fill = hotel_principal_fondo
                    cell.font = hotel_principal_fuente
                elif "Tarifa promedio de competidores" in str(hotel_name) or "Disponibilidad de la oferta" in str(hotel_name):
                    cell.fill = metricas_fondo
                    cell.font = metricas_fuente
                elif "Diferencia de mi tarifa" in str(hotel_name):
                    cell.fill = metricas_fondo
                    cell.font = metricas_fuente
                    # Formato condicional para diferencias
                    if col_idx > 2 and value and str(value).replace('.', '').replace(',', '').replace('-', '').isdigit():
                        try:
                            diff_value = float(str(value).replace(',', '.'))
                            if diff_value < 0:
                                cell.font = verde_fuente
                            elif diff_value > 0:
                                cell.font = rojo_fuente
                        except (ValueError, TypeError):
                            pass
                else:  # Competidores
                    cell.font = competidor_fuente
        
        # 3. FORMATEAR ENCABEZADOS (Fila 2)
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.fill = encabezado_fondo
            cell.font = encabezado_fuente
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = borde_gris
            
            # Borde inferior blanco para encabezados
            cell.border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thick', color='FFFFFF')
            )
        
        # 4. AJUSTAR ANCHO DE COLUMNAS
        for col_idx, col_name in enumerate(df.columns, start=1):
            column_letter = get_column_letter(col_idx)
            if col_name == "Hotel Name":
                ws.column_dimensions[column_letter].width = 25
            elif col_name == "URL":
                ws.column_dimensions[column_letter].width = 50
            else:
                ws.column_dimensions[column_letter].width = 15
        
        # 5. CONGELAR PANELES (Título y encabezados)
        ws.freeze_panes = 'A3'
        
        # 6. ALINEACIÓN DE TEXTO
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for col_idx, cell in enumerate(row, start=1):
                if col_idx <= 2:  # Hotel Name y URL
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:  # Fechas y precios
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # 7. AGREGAR PIE DE PÁGINA
        # Agregar fila vacía después de los datos
        ws.insert_rows(ws.max_row + 1)
        
        # Agregar fila de pie de página
        pie_fila = ws.max_row + 1
        pie_cell = ws.cell(row=pie_fila, column=1, value="Con tecnología de www.HotelRateShopper.com")
        pie_cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        pie_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Pintar fondo azul en toda la fila del pie de página
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=pie_fila, column=col_idx)
            cell.fill = titulo_fondo  # Mismo fondo azul que el título
        
        # 8. GUARDAR EL ARCHIVO CON FORMATO APLICADO
        wb.save('temp_excel_download.xlsx')
        
        with open('temp_excel_download.xlsx', 'rb') as f:
            excel_bytes = f.read()
        os.remove('temp_excel_download.xlsx')
        # Generar nombre de archivo con formato HotelRateShopper_YYMMDD_NombreSetCompetitivo
        fecha_formato = datetime.now().strftime('%y%m%d')
        nombre_archivo = f"HotelRateShopper_{fecha_formato}_{report_data.get('setName', 'Reporte')}.xlsx"
        
        return send_file(
            io.BytesIO(excel_bytes),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nombre_archivo
        )
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/scraper-status', methods=['GET'])
def get_scraper_status():
    return jsonify(scraper_status)

# --- FUNCIÓN DE CONEXIÓN A APIS EXTERNAS ---
def obtener_datos_externos():
    try:
        # Conectar a tu API externa aquí
        headers = {
            'Authorization': 'Bearer TU_TOKEN_AQUI',
            'Content-Type': 'application/json'
        }
        response = requests.get('TU_API_URL_AQUI', headers=headers)
        if response.status_code == 200:
            return response.json()
        else:
            raise Exception(f"Error en API: {response.status_code}")
    except Exception as e:
        # Datos de ejemplo para testing
        return [
            {"id": 1, "nombre": "Dato de prueba 1", "fecha": datetime.now().isoformat()},
            {"id": 2, "nombre": "Dato de prueba 2", "fecha": datetime.now().isoformat()}
        ]

# --- MANTENER EL RESTO DE ENDPOINTS ---
@app.route('/execute-scheduled-tasks', methods=['POST'])
def execute_scheduled_tasks():
    try:
        # Validar token de seguridad (opcional pero recomendado)
        token = request.args.get('token')
        expected_token = os.environ.get('SCHEDULER_TOKEN', 'default_token')
        
        if token != expected_token:
            logger.warning(f"[execute-scheduled-tasks] Token inválido recibido: {token}")
            return jsonify({"error": "Token inválido"}), 401
        
        logger.info("[execute-scheduled-tasks] Iniciando ejecución de tareas programadas")
        
        # Obtener todos los usuarios que tienen grupos con schedule_enabled = True
        users_ref = db.collection('users')
        users = users_ref.stream()
        
        total_tasks_created = 0
        
        for user_doc in users:
            uid = user_doc.id
            user_data = user_doc.to_dict()
            
            # Obtener grupos del usuario
            grupos_ref = db.collection('users').document(uid).collection('grupos')
            grupos = grupos_ref.stream()
            
            for grupo in grupos:
                grupo_data = grupo.to_dict()
                grupo_id = grupo.id
                
                # Verificar si el grupo tiene schedule habilitado
                if grupo_data.get('schedule_enabled', False):
                    logger.info(f"[execute-scheduled-tasks] Procesando grupo {grupo_id} del usuario {uid}")
                    
                    # Obtener URLs de hoteles
                    hotel_principal = grupo_data.get('hotel_principal')
                    competidores = grupo_data.get('competidores', [])
                    
                    if not hotel_principal:
                        logger.warning(f"[execute-scheduled-tasks] Grupo {grupo_id} no tiene hotel principal")
                        continue
                    
                    # Construir lista de URLs
                    hotel_urls = [hotel_principal]
                    if competidores:
                        if isinstance(competidores, list):
                            hotel_urls.extend(competidores)
                        elif isinstance(competidores, str):
                            hotel_urls.append(competidores)
                    
                    # Crear documento de reporte
                    report_doc = {
                        'userId': uid,
                        'setId': grupo_id,
                        'setName': grupo_data.get('name', 'Grupo Programado'),
                        'ownHotelUrl': hotel_principal,
                        'competitorHotelUrls': competidores,
                        'hotel_base_urls': hotel_urls,
                        'days': grupo_data.get('days', 7),
                        'nights': grupo_data.get('nights', 1),
                        'currency': grupo_data.get('currency', 'USD'),
                        'userEmail': user_data.get('email'),
                        'status': 'queued',
                        'createdAt': datetime.now(),
                        'scheduled_task': True
                    }
                    
                    # Guardar en Firestore
                    try:
                        db.collection('scraping_reports').add(report_doc)
                        total_tasks_created += 1
                        logger.info(f"[execute-scheduled-tasks] ✅ Tarea creada para grupo {grupo_id} del usuario {uid}")
                    except Exception as e:
                        logger.error(f"[execute-scheduled-tasks] ❌ Error creando tarea para grupo {grupo_id}: {e}")
        
        logger.info(f"[execute-scheduled-tasks] ✅ Proceso completado. {total_tasks_created} tareas creadas")
        return jsonify({
            "success": True, 
            "message": f"Tareas programadas encoladas: {total_tasks_created}",
            "tasks_created": total_tasks_created
        })
        
    except Exception as e:
        logger.error(f"[execute-scheduled-tasks] ❌ Error general: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/init-user', methods=['POST'])
def init_user():
    try:
        data = request.get_json()
        uid = data.get('uid')
        email = data.get('email')
        plan = data.get('plan', 'free_trial')
        if not uid or not email:
            return jsonify({"error": "UID y email requeridos"}), 400
        # Crear o actualizar usuario
        user_ref = db.collection('users').document(uid)
        user_ref.set({
            'email': email,
            'plan': plan,
            'created_at': datetime.now()
        })
        return jsonify({"success": True, "message": "Usuario inicializado"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/crear-grupo', methods=['POST'])
def crear_grupo():
    try:
        data = request.get_json()
        uid = data.get('uid')
        nombre = data.get('nombre')
        hotel_principal = data.get('hotel_principal')
        if not uid or not nombre or not hotel_principal:
            return jsonify({"error": "UID, nombre y hotel principal requeridos"}), 400
        # Verificar límites del plan
        user_plan = get_user_plan(uid)
        plan_limits = PLAN_LIMITS.get(user_plan, PLAN_LIMITS["free_trial"])
        # Contar grupos existentes
        grupos_ref = db.collection('users').document(uid).collection('grupos')
        grupos_existentes = len(list(grupos_ref.stream()))
        if grupos_existentes >= plan_limits["max_groups"]:
            return jsonify({
                "error": f"Tu plan {user_plan} permite máximo {plan_limits['max_groups']} grupos"
            }), 400
        # Crear grupo
        grupo_ref = db.collection('users').document(uid).collection('grupos').document()
        grupo_ref.set({
            'name': nombre,
            'hotel_principal': hotel_principal,
            'competidores': [],
            'days': 7,
            'schedule_enabled': False,
            'created_at': datetime.now()
        })
        return jsonify({"success": True, "message": "Grupo creado"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/agregar-competidor', methods=['POST'])
def agregar_competidor():
    try:
        data = request.get_json()
        uid = data.get('uid')
        grupo_id = data.get('grupo_id')
        competidor_url = data.get('competidor_url')
        if not uid or not grupo_id or not competidor_url:
            return jsonify({"error": "UID, grupo_id y competidor_url requeridos"}), 400
        # Verificar límites del plan
        user_plan = get_user_plan(uid)
        plan_limits = PLAN_LIMITS.get(user_plan, PLAN_LIMITS["free_trial"])
        # Obtener grupo actual
        grupo_ref = db.collection('users').document(uid).collection('grupos').document(grupo_id)
        grupo = grupo_ref.get()
        if not grupo.exists:
            return jsonify({"error": "Grupo no encontrado"}), 404
        grupo_data = grupo.to_dict()
        competidores_actuales = len(grupo_data.get('competidores', []))
        if competidores_actuales >= plan_limits["max_competitors"]:
            return jsonify({
                "error": f"Tu plan {user_plan} permite máximo {plan_limits['max_competitors']} competidores"
            }), 400
        # Agregar competidor
        competidores = grupo_data.get('competidores', [])
        competidores.append(competidor_url)
        grupo_ref.update({
            'competidores': competidores
        })
        return jsonify({"success": True, "message": "Competidor agregado"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/configurar-dias', methods=['POST'])
def configurar_dias():
    try:
        data = request.get_json()
        uid = data.get('uid')
        grupo_id = data.get('grupo_id')
        dias = data.get('dias')
        if not uid or not grupo_id or not dias:
            return jsonify({"error": "UID, grupo_id y dias requeridos"}), 400
        # Verificar límites del plan
        user_plan = get_user_plan(uid)
        plan_limits = PLAN_LIMITS.get(user_plan, PLAN_LIMITS["free_trial"])
        if dias > plan_limits["max_days"]:
            return jsonify({
                "error": f"Tu plan {user_plan} permite máximo {plan_limits['max_days']} días"
            }), 400
        # Actualizar días
        grupo_ref = db.collection('users').document(uid).collection('grupos').document(grupo_id)
        grupo_ref.update({
            'days': dias
        })
        return jsonify({"success": True, "message": "Días configurados"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/configurar-schedule', methods=['POST'])
def configurar_schedule():
    try:
        data = request.get_json()
        uid = data.get('uid')
        grupo_id = data.get('grupo_id')
        enabled = data.get('enabled', False)
        if not uid or not grupo_id:
            return jsonify({"error": "UID y grupo_id requeridos"}), 400
        # Verificar si el plan permite scheduling
        user_plan = get_user_plan(uid)
        plan_limits = PLAN_LIMITS.get(user_plan, PLAN_LIMITS["free_trial"])
        if enabled and not plan_limits["scheduling"]:
            return jsonify({
                "error": f"Tu plan {user_plan} no permite programación automática"
            }), 400
        # Actualizar schedule
        grupo_ref = db.collection('users').document(uid).collection('grupos').document(grupo_id)
        grupo_ref.update({
            'schedule_enabled': enabled
        })
        return jsonify({"success": True, "message": "Schedule configurado"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/create-subscription-checkout', methods=['POST'])
def create_subscription_checkout():
    try:
        data = request.get_json()
        plan = data.get('plan')
        user_email = data.get('user_email')
        if not plan or not user_email:
            return jsonify({"error": "Plan y email requeridos"}), 400
        if plan not in MP_PLAN_IDS:
            return jsonify({"error": "Plan no válido"}), 400
        # Configurar Mercado Pago
        mp = mercadopago.SDK(os.environ.get('MP_ACCESS_TOKEN'))
        # Crear preferencia de suscripción
        preference_data = {
            "items": [
                {
                    "title": f"Plan {plan}",
                    "quantity": 1,
                    "unit_price": 10.0  # Precio de ejemplo
                }
            ],
            "payer": {
                "email": user_email
            },
            "back_urls": {
                "success": "https://tu-frontend.com/success",
                "failure": "https://tu-frontend.com/failure",
                "pending": "https://tu-frontend.com/pending"
            },
            "auto_return": "approved",
            "external_reference": f"plan_{plan}_{user_email}",
            "notification_url": "https://tu-backend.com/mercado-pago-webhook"
        }
        preference = mp.preference().create(preference_data)
        if preference["status"] == "created":
            return jsonify({
                "success": True,
                "init_point": preference["response"]["init_point"],
                "preference_id": preference["response"]["id"]
            })
        else:
            return jsonify({"error": "Error creando preferencia"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/mercado-pago-webhook', methods=['POST'])
def mercado_pago_webhook():
    try:
        data = request.get_json()
        logger.info(f"Webhook recibido: {data}")
        
        # Verificar que es una notificación válida de MP
        notification_type = data.get("type")
        
        if notification_type == "payment":
            # Manejar notificaciones de pagos únicos
            payment_id = data.get("data", {}).get("id")
            
            if payment_id:
                # Obtener información del pago
                mp = mercadopago.SDK(os.environ.get('MP_ACCESS_TOKEN'))
                payment_info = mp.payment().get(payment_id)
                
                if payment_info["status"] == 200:
                    payment_data = payment_info["response"]
                    
                    # Procesar según el estado del pago
                    if payment_data["status"] == "approved":
                        # Pago aprobado - actualizar plan del usuario
                        external_reference = payment_data.get("external_reference", "")
                        
                        # Extraer información del external_reference
                        # Formato esperado: "plan_{plan}_{email}"
                        if external_reference.startswith("plan_"):
                            parts = external_reference.split("_")
                            if len(parts) >= 3:
                                plan = parts[1]
                                email = "_".join(parts[2:])  # En caso de que el email tenga _
                                
                                # Buscar usuario por email
                                users_ref = db.collection('users')
                                users = users_ref.where('email', '==', email).stream()
                                
                                for user in users:
                                    user_ref = db.collection('users').document(user.id)
                                    user_ref.update({
                                        'plan': plan,
                                        'plan_updated_at': datetime.now()
                                    })
                                    
                                    # Enviar email de confirmación
                                    try:
                                        from mailersend import emails
                                        mailer = emails.NewEmail(os.environ.get('MAILERSEND_API_KEY'))
                                        
                                        mail_body = {}
                                        
                                        mail_from = {
                                            "email": "noreply@tuapp.com",
                                            "name": "Tu App"
                                        }
                                        
                                        recipients = [
                                            {
                                                "email": email,
                                                "name": "Usuario"
                                            }
                                        ]
                                        
                                        mailer.set_mail_from(mail_from, mail_body)
                                        mailer.set_mail_to(recipients, mail_body)
                                        mailer.set_subject("Plan actualizado exitosamente", mail_body)
                                        mailer.set_html_content(f"""
                                        <h2>¡Plan actualizado!</h2>
                                        <p>Tu plan ha sido actualizado a <strong>{plan}</strong>.</p>
                                        <p>Gracias por tu compra.</p>
                                        """, mail_body)
                                        
                                        mailer.send(mail_body)
                                        
                                    except Exception as e:
                                        logger.error(f"Error enviando email: {e}")
                                    
                                    break
                
                    elif payment_data["status"] == "rejected":
                        # Pago rechazado - enviar email de notificación
                        external_reference = payment_data.get("external_reference", "")
                        if external_reference.startswith("plan_"):
                            parts = external_reference.split("_")
                            if len(parts) >= 3:
                                email = "_".join(parts[2:])
                                
                                try:
                                    from mailersend import emails
                                    mailer = emails.NewEmail(os.environ.get('MAILERSEND_API_KEY'))
                                    
                                    mail_body = {}
                                    
                                    mail_from = {
                                        "email": "noreply@tuapp.com",
                                        "name": "Tu App"
                                    }
                                    
                                    recipients = [
                                        {
                                            "email": email,
                                            "name": "Usuario"
                                        }
                                    ]
                                    
                                    mailer.set_mail_from(mail_from, mail_body)
                                    mailer.set_mail_to(recipients, mail_body)
                                    mailer.set_subject("Problema con el pago", mail_body)
                                    mailer.set_html_content("""
                                    <h2>Problema con el pago</h2>
                                    <p>Tu pago fue rechazado. Por favor, intenta nuevamente.</p>
                                    """, mail_body)
                                    
                                    mailer.send(mail_body)
                                    
                                except Exception as e:
                                    logger.error(f"Error enviando email: {e}")
        
        elif notification_type == "preapproval":
            # Manejar notificaciones de suscripciones (preapproval)
            preapproval_id = data.get("data", {}).get("id")
            
            if preapproval_id:
                # Obtener información de la suscripción
                mp = mercadopago.SDK(os.environ.get('MP_ACCESS_TOKEN'))
                preapproval_info = mp.preapproval().get(preapproval_id)
                
                if preapproval_info["status"] == 200:
                    preapproval_data = preapproval_info["response"]
                    
                    # Procesar según el estado de la suscripción
                    if preapproval_data["status"] == "authorized":
                        # Suscripción autorizada - actualizar plan del usuario
                        external_reference = preapproval_data.get("external_reference", "")
                        payer_email = preapproval_data.get("payer_email", "")
                        
                        # Determinar el plan basado en el preapproval_plan_id
                        plan_id = preapproval_data.get("preapproval_plan_id", "")
                        plan_mapping = {
                            "2c93808497c462520197d744586508be": "esencial",
                            "2c93808497c19ac40197d7445b440a20": "pro", 
                            "2c93808497d635430197d7445e1c00bc": "market_leader"
                        }
                        plan = plan_mapping.get(plan_id, "esencial")
                        
                        # Buscar usuario por email o external_reference
                        users_ref = db.collection('users')
                        users = users_ref.where('email', '==', payer_email).stream()
                        
                        user_found = False
                        for user in users:
                            user_ref = db.collection('users').document(user.id)
                            user_ref.update({
                                'plan': plan,
                                'plan_updated_at': datetime.now(),
                                'subscription_id': preapproval_id,
                                'subscription_status': preapproval_data["status"]
                            })
                            
                            # Enviar email de confirmación
                            try:
                                from mailersend import emails
                                mailer = emails.NewEmail(os.environ.get('MAILERSEND_API_KEY'))
                                
                                mail_body = {}
                                
                                mail_from = {
                                    "email": "noreply@tuapp.com",
                                    "name": "Tu App"
                                }
                                
                                recipients = [
                                    {
                                        "email": payer_email,
                                        "name": "Usuario"
                                    }
                                ]
                                
                                mailer.set_mail_from(mail_from, mail_body)
                                mailer.set_mail_to(recipients, mail_body)
                                mailer.set_subject("Suscripción activada exitosamente", mail_body)
                                mailer.set_html_content(f"""
                                <h2>¡Suscripción activada!</h2>
                                <p>Tu suscripción al plan <strong>{plan}</strong> ha sido activada exitosamente.</p>
                                <p>Gracias por tu compra.</p>
                                """, mail_body)
                                
                                mailer.send(mail_body)
                                
                            except Exception as e:
                                logger.error(f"Error enviando email: {e}")
                            
                            user_found = True
                            break
                        
                        if not user_found:
                            logger.warning(f"Usuario no encontrado para email: {payer_email}")
                    
                    elif preapproval_data["status"] == "rejected":
                        # Suscripción rechazada - enviar email de notificación
                        payer_email = preapproval_data.get("payer_email", "")
                        
                        try:
                            from mailersend import emails
                            mailer = emails.NewEmail(os.environ.get('MAILERSEND_API_KEY'))
                            
                            mail_body = {}
                            
                            mail_from = {
                                "email": "noreply@tuapp.com",
                                "name": "Tu App"
                            }
                            
                            recipients = [
                                {
                                    "email": payer_email,
                                    "name": "Usuario"
                                }
                            ]
                            
                            mailer.set_mail_from(mail_from, mail_body)
                            mailer.set_mail_to(recipients, mail_body)
                            mailer.set_subject("Problema con la suscripción", mail_body)
                            mailer.set_html_content("""
                            <h2>Problema con la suscripción</h2>
                            <p>Tu suscripción fue rechazada. Por favor, intenta nuevamente.</p>
                            """, mail_body)
                            
                            mailer.send(mail_body)
                            
                        except Exception as e:
                            logger.error(f"Error enviando email: {e}")
        
        return jsonify({"success": True}), 200
        
    except Exception as e:
        logger.error(f"Error en webhook de Mercado Pago: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/debug', methods=['POST'])
def debug_endpoint():
    try:
        data = request.get_json()
        uid = data.get('uid')
        
        if not uid:
            return jsonify({"error": "UID requerido"}), 400
        
        # Obtener información del usuario
        user_ref = db.collection('users').document(uid)
        user_doc = user_ref.get()
        
        if not user_doc.exists:
            return jsonify({"error": "Usuario no encontrado"}), 404
        
        user_data = user_doc.to_dict()
        
        # Obtener grupos del usuario
        grupos_ref = db.collection('users').document(uid).collection('grupos')
        grupos = []
        for grupo in grupos_ref.stream():
            grupos.append({
                'id': grupo.id,
                'data': grupo.to_dict()
            })
        
        return jsonify({
            "user": user_data,
            "grupos": grupos,
            "scraper_status": scraper_status
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/test-metricas', methods=['GET'])
def test_metricas():
    return jsonify({"message": "Endpoint de test de métricas"})

@app.route('/debug-update-task', methods=['POST'])
def debug_update_task():
    """
    Endpoint temporal para probar actualización manual en Firestore
    """
    try:
        data = request.get_json()
        report_id = data.get('report_id')
        new_status = data.get('status', 'completed')
        
        if not report_id:
            return jsonify({"success": False, "error": "report_id es requerido"}), 400
        
        logger.info(f"[DEBUG] Intentando actualizar tarea {report_id} a status: {new_status}")
        
        # Intentar actualizar el documento
        try:
            now = datetime.now()
            update_data = {
                "status": new_status,
                "completedAt": now,
                "debug_updated_at": now
            }
            
            logger.info(f"[DEBUG] Datos a actualizar: {update_data}")
            
            doc_ref = db.collection("scraping_reports").document(report_id)
            
            # Verificar si el documento existe
            doc = doc_ref.get()
            if not doc.exists:
                logger.error(f"[DEBUG] ❌ Documento {report_id} no existe en Firestore")
                return jsonify({"success": False, "error": f"Documento {report_id} no existe"}), 404
            
            logger.info(f"[DEBUG] ✅ Documento {report_id} existe. Actualizando...")
            
            # Intentar actualizar
            doc_ref.update(update_data)
            
            logger.info(f"[DEBUG] ✅ Documento {report_id} actualizado exitosamente a status: {new_status}")
            
            return jsonify({
                "success": True, 
                "message": f"Documento {report_id} actualizado a {new_status}",
                "report_id": report_id,
                "new_status": new_status
            })
            
        except Exception as e:
            logger.error(f"[DEBUG] ❌ Error actualizando documento {report_id}: {e}")
            logger.error(f"[DEBUG] ❌ Tipo de error: {type(e)}")
            return jsonify({
                "success": False, 
                "error": f"Error actualizando Firestore: {str(e)}",
                "error_type": str(type(e))
            }), 500
            
    except Exception as e:
        logger.error(f"[DEBUG] ❌ Error general en debug-update-task: {e}")
        return jsonify({"success": False, "error": f"Error general: {str(e)}"}), 500

@app.route('/test-scheduled-tasks', methods=['GET'])
def test_scheduled_tasks():
    """
    Endpoint de prueba para verificar el estado del sistema de tareas programadas
    """
    try:
        logger.info("[test-scheduled-tasks] Verificando estado del sistema")
        
        # Contar usuarios con grupos programados
        users_with_scheduled_groups = 0
        total_scheduled_groups = 0
        
        users_ref = db.collection('users')
        users = users_ref.stream()
        
        for user_doc in users:
            uid = user_doc.id
            user_data = user_doc.to_dict()
            
            # Obtener grupos del usuario
            grupos_ref = db.collection('users').document(uid).collection('grupos')
            grupos = grupos_ref.stream()
            
            user_has_scheduled = False
            for grupo in grupos:
                grupo_data = grupo.to_dict()
                if grupo_data.get('schedule_enabled', False):
                    user_has_scheduled = True
                    total_scheduled_groups += 1
            
            if user_has_scheduled:
                users_with_scheduled_groups += 1
        
        # Contar tareas en cola
        queued_tasks = list(db.collection('scraping_reports').where('status', '==', 'queued').stream())
        pending_tasks = list(db.collection('scraping_reports').where('status', '==', 'pending').stream())
        
        return jsonify({
            "success": True,
            "system_status": {
                "users_with_scheduled_groups": users_with_scheduled_groups,
                "total_scheduled_groups": total_scheduled_groups,
                "queued_tasks": len(queued_tasks),
                "pending_tasks": len(pending_tasks),
                "scraper_running": scraper_status["is_running"],
                "scraper_en_proceso": scraper_en_proceso.is_set()
            },
            "message": "Estado del sistema verificado"
        })
        
    except Exception as e:
        logger.error(f"[test-scheduled-tasks] ❌ Error: {e}")
        return jsonify({"error": str(e)}), 500

# Lanzar el procesador de la cola SIEMPRE, no solo en modo script
import threading
threading.Thread(target=cola_procesadora_scraping, daemon=True).start()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000) 